#%%
###import packages

#
import streamlit as st
import streamlit.components.v1 as components
from transformers import pipeline
import pandas as pd
import time
import openpyxl  
import base64   

#%%
st.set_page_config(layout="wide")


# Text/Title

with st.container():
    components.html(
     """
     <div style="position: fixed;
   left: 0;
   bottom: 0;
   width: 100%;
   color: black;
   
   text-align: center;">
  <h1>Sentiment Analysis</h1>
</div>
     """,height=50,)

#%%
#Navigation bar
with st.container():
 #navbar 
#https://bootsnipp.com/snippets/nNX3a     https://www.mockplus.com/blog/post/bootstrap-navbar-template
   components.html(
       """
       <link href="//maxcdn.bootstrapcdn.com/bootstrap/4.1.1/css/bootstrap.min.css" rel="stylesheet" id="bootstrap-css">
<script src="//maxcdn.bootstrapcdn.com/bootstrap/4.1.1/js/bootstrap.min.js"></script>
<script src="//cdnjs.cloudflare.com/ajax/libs/jquery/3.2.1/jquery.min.js"></script>
<!------ Include the above in your HEAD tag ---------->
<nav class="navbar navbar-icon-top navbar-expand-lg navbar-dark bg-dark" >
  <a class="navbar-brand" href="https://www.rstiwari.com" target="_blank">Profile</a>
  <button class="navbar-toggler" type="button" data-toggle="collapse" data-target="#navbarSupportedContent" aria-controls="navbarSupportedContent" aria-expanded="false" aria-label="Toggle navigation">
    <span class="navbar-toggler-icon"></span>
  </button>
  <div class="collapse navbar-collapse" id="navbarSupportedContent">
    <ul class="navbar-nav mr-auto">
      <li class="nav-item active">
        <a class="nav-link" href=" https://tiwari11-rst.medium.com/" target="_blank" >
          <i class="fa fa-home"></i>
          Medium
          <span class="sr-only">(current)</span>
          </a>
      </li>
      <li class="nav-item">
        <a class="nav-link" href=" https://happyman11.github.io/" target="_blank">
          <i class="fa fa-envelope-o">
            <span class="badge badge-danger">Git Pages</span>
          </i>
         
        </a>
      </li>
      
        <li class="nav-item">
        <a class="nav-link" href="https://happyman11.github.io/" target="_blank">
          <i class="fa fa-globe">
            <span class="badge badge-success">Badges</span>
          </i>
         
        </a>
      </li>
          
        </a>
      </li>
      
      <li class="nav-item">
        <a class="nav-link disabled" href="https://ravishekhartiwari.blogspot.com/" target="_blank">
          <i class="fa fa-envelope-o">
            <span class="badge badge-warning">Blogspot</span>
          </i>
          
        </a>
      </li>
      
      
    </ul>
  
    
  </div>
</nav>
       """, height=70,
    )



###
classifier = pipeline("sentiment-analysis")
def analysis_sentence(comments):

    output=classifier([comments])
    return (output)
    
def format_output(ouput):
    formatted_output={ "label": output[0]["label"],
                        "Probability":output[0]["score"]
                        }
                        
    return(formatted_output) 
    
def read_pdf(path):
    
    if path is not None:
     
        List_email = openpyxl.load_workbook(path)
        
        return(List_email)
        

def filedownload(df):
    csv = df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()  # strings <-> bytes conversions
    href = f'<a href="data:file/csv;base64,{b64}" download="Analysed Comments.csv">Download CSV File</a>'
    return href

st.sidebar.header("Bulk Analysis ") 
attachment_comments = st.sidebar.file_uploader("Upload file to be analysed", type=["xls" , "xlsx" , "xlsm" , "xlsb" , "odf" , "ods","odt"])
Submitted_file=st.sidebar.button("Sentiments")

with st.container():
     
     col10, col11, col12 = st.columns([2,4,3])   
     with col11:
         sentiments= st.text_input("Sentence","Type  here...")



with st.container():
     
     col10, col11, col12 = st.columns(3)   
     with col11:
         Submitted=st.button("Process")

if (Submitted):
    output=analysis_sentence(sentiments)
    with st.spinner('Formatting output '):
     time.sleep(5)
    analysis=format_output(output)
    
    
    st.header("Prediction from model")
    st.write(pd.DataFrame({
                           "Comments":[sentiments],
                            "Label" :[analysis["label"]],
                            "Probabilities":[analysis["Probability"]],
                          }))




##add for excel

if(Submitted_file):
   if attachment_comments is not None:
       extracted_comments = read_pdf(attachment_comments)
       sheet_obj = extracted_comments.active 
       m_row = sheet_obj.max_row 
       #st.write(m_row)
       labels=[]
       probability=[]
       comments_extracted=[]
       prog=0
       my_bar = st.progress(0)
       jump=100/m_row 
       for i in range(1, m_row + 1):
           msg='Processing'+ str(i) +"out of "+ str(m_row)
          
           
           cell_obj = sheet_obj.cell(row = i, column = 1)
           comments=cell_obj.value
           output=analysis_sentence(comments)
           analysis=format_output(output)
           
           
           comments_extracted.append(comments)
           labels.append(analysis["label"])
           probability.append(analysis["Probability"])
           
           msg=""
           msg='Processed'+ str(i) +"out of "+ str(m_row)
           my_bar.progress(prog + jump)
           
           st.spinner(msg) 
           time.sleep(2)
           msg=""


       st.header("Prediction from model")
       data_file=pd.DataFrame({
                           "Comments":comments_extracted,
                            "Label" :labels,
                            "Probabilities":probability,
                          })
       st.write(pd.DataFrame({
                           "Comments":comments_extracted,
                            "Label" :labels,
                            "Probabilities":probability,
                          }))
       st.markdown(filedownload(data_file), unsafe_allow_html=True)

    
 
#footer

with st.container():
    components.html(
     """
     <div style="position: fixed;
   left: 0;
   bottom: 0;
   width: 100%;
   background-color: black;
   color: white;
   text-align: center;">
  <p>Ravi Shekhar Tiwari</p>
</div>
     """,height=140,)